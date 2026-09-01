import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from scripts.build_wzdb import build_database, event_component, event_date_assignment_stats
from scripts.event_dates import event_mapping_key, load_event_date_mapping
from scripts.match_event_dates import (
    EVENT_FIELDS,
    LogicalEvent,
    PhysicalEvent,
    apply_reviewed_supplement,
    decode_events,
    match_events,
    venue_key,
)


def make_event(index=0, fragments=None, teams=None, **overrides):
    values = {
        "home": "",
        "away": "",
        "score": "",
        "league": "International",
        "track": "Łódź",
        "competition": "GP",
        "round": "7 runda",
        "capacity": "",
    }
    values.update(overrides)
    fragments = fragments or [values]
    physical = []
    start = index * 20
    for offset, fragment_values in enumerate(fragments):
        signature = tuple(fragment_values[field] for field in EVENT_FIELDS)
        physical.append(
            PhysicalEvent(
                source_event_index=index + offset,
                start=start + offset * 5,
                count=5,
                values=fragment_values,
                mapping_key=event_mapping_key("2026", signature, 0),
            )
        )
    return LogicalEvent(index, physical, values, teams or [])


def candidate(date="2026-08-01", venue="Łódź", event="FIM SGP of Poland - Lodz - Runda 7", competition="FIM SGP", **overrides):
    result = {
        "date": date,
        "venue": venue,
        "event": event,
        "competition": competition,
        "source": "test calendar",
        "source_detail": "fixture",
        "source_page": 1,
        "confidence": "HIGH",
        "match_key": f"{date}|{venue}|{event}",
    }
    result.update(overrides)
    return result


class LogicalEventDecodingTests(unittest.TestCase):
    def decode_fragments(self, fragments, counts=None, dates=None, event_teams=None):
        strings = [""]

        def intern(value):
            value = str(value or "")
            if value not in strings:
                strings.append(value)
            return strings.index(value)

        counts = counts or [1] * len(fragments)
        dates = dates or ["2026-08-30"] * len(fragments)
        event_teams = event_teams or [[] for _ in fragments]
        rows = []
        events = []
        start = 0
        for fragment, count, event_date, teams in zip(
            fragments, counts, dates, event_teams
        ):
            row = [0, 0, 0, 0, 0]
            row.extend(intern(fragment.get(field, "")) for field in EVENT_FIELDS)
            rows.extend([row] * count)
            events.append([start, count, 1, teams, intern(event_date)])
            start += count
        return decode_events(
            {
                "strings": strings,
                "years": {"2026": rows},
                "events": {"2026": events},
            },
            "2026",
        )

    def decode_two_team_fragments(self, capacities):
        strings = [""]

        def intern(value):
            value = str(value or "")
            if value not in strings:
                strings.append(value)
            return strings.index(value)

        rows = []
        for team, score, capacity in zip(
            ("Team A", "Team B"), ("30", "20"), capacities
        ):
            rows.append([
                0, 0, 0, 0, 0,
                intern(team), 0, intern(score), intern("Test league"),
                intern("Test track"), intern("Team competition"),
                intern("1 runda"), intern(capacity),
            ])
        date_id = intern("2026-05-01")
        return decode_events({
            "strings": strings,
            "years": {"2026": rows},
            "events": {"2026": [
                [0, 1, 1, [], date_id],
                [1, 1, 1, [], date_id],
            ]},
        }, "2026")

    def test_albury_rounds_keep_mapping_fragments_but_decode_to_two_events(self):
        strings = [""]
        string_ids = {"": 0}

        def intern(value):
            value = str(value or "")
            if value not in string_ids:
                string_ids[value] = len(strings)
                strings.append(value)
            return string_ids[value]

        def record(round_name, home="", away="", score="", capacity=""):
            return [
                0, 0, 0, 0, 0,
                intern(home), intern(away), intern(score), intern("Australia"),
                intern("Albury-Wodonga"), intern("IM Australii"),
                intern(round_name), intern(capacity),
            ]

        rows = [record("1 runda") for _ in range(17)]
        rows.extend([
            record("2 runda"),
            record("2 runda", "2", "miejsce w finale"),
            record("2 runda", "3", "miejsce w finale"),
            *[record("2 runda") for _ in range(14)],
        ])
        first_date = intern("2026-01-03")
        second_date = intern("2026-01-04")
        database = {
            "strings": strings,
            "years": {"2026": rows},
            "events": {"2026": [
                [0, 17, 1, [], first_date],
                [17, 1, 1, [], second_date],
                [18, 1, 1, [], second_date],
                [19, 1, 1, [], second_date],
                [20, 14, 1, [], second_date],
            ]},
        }

        logical = decode_events(database, "2026")

        self.assertEqual(len(database["events"]["2026"]), 5)
        self.assertEqual(len(logical), 2)
        self.assertEqual(
            [(event.start, event.count, event.values["round"]) for event in logical],
            [(0, 17, "1 runda"), (17, 17, "2 runda")],
        )
        self.assertEqual(
            [[fragment.event_date for fragment in event.physical] for event in logical],
            [["2026-01-03"], ["2026-01-04"] * 4],
        )

    def test_dated_individual_categories_with_different_capacity_stay_separate(self):
        database = {
            "strings": ["", "Miniżużel", "Glasgow", "British Youth Championship", "1 runda", "125 ccm", "250 ccm", "2026-04-12"],
            "years": {"2026": [
                [0, 0, 0, 0, 0, 0, 0, 0, 1, 2, 3, 4, 5],
                [0, 0, 0, 0, 0, 0, 0, 0, 1, 2, 3, 4, 6],
            ]},
            "events": {"2026": [
                [0, 1, 1, [], 7],
                [1, 1, 1, [], 7],
            ]},
        }

        logical = decode_events(database, "2026")

        self.assertEqual(len(logical), 2)

    def test_multi_team_fragments_with_different_capacity_stay_separate(self):
        logical = self.decode_two_team_fragments(("125 ccm", "250 ccm"))

        self.assertEqual(len(logical), 2)

    def test_multi_team_fragments_with_identical_capacity_merge(self):
        logical = self.decode_two_team_fragments(("500R", "500R"))

        self.assertEqual(len(logical), 1)
        self.assertEqual(len(logical[0].teams), 2)

    def test_multi_team_fragments_with_both_capacities_blank_merge(self):
        logical = self.decode_two_team_fragments(("", ""))

        self.assertEqual(len(logical), 1)
        self.assertEqual(len(logical[0].teams), 2)

    def test_multi_team_fragments_with_blank_and_nonblank_capacity_stay_separate(self):
        logical = self.decode_two_team_fragments(("", "500R"))

        self.assertEqual(len(logical), 2)

    def test_zarnovica_rider_specific_g_i_merges_to_one_logical_event(self):
        base = {
            "league": "Słowacja",
            "track": "Żarnowica",
            "competition": "Zlata Prilba",
            "round": "",
            "capacity": "",
        }
        fragments = [
            {**base, "home": str(place), "away": "", "score": "miejsce w finale"}
            for place in range(1, 10)
        ]
        fragments.append({**base, "home": "", "away": "", "score": ""})

        logical = self.decode_fragments(fragments, counts=[1] * 9 + [11])

        self.assertEqual(len(logical), 1)
        self.assertEqual(logical[0].count, 20)
        self.assertEqual(len(logical[0].physical), 10)
        self.assertEqual(logical[0].fragment_count, 10)
        self.assertEqual(logical[0].teams, [])

    def test_event_teams_alone_do_not_make_rider_fragments_team_shaped(self):
        base = {
            "league": "Słowacja",
            "track": "Żarnowica",
            "competition": "Zlata Prilba",
            "round": "",
            "capacity": "",
        }
        logical = self.decode_fragments(
            [
                {**base, "home": "1", "score": "miejsce w finale"},
                {**base, "home": "2", "score": "miejsce w finale"},
            ],
            event_teams=[
                [{"name": "1", "score": "miejsce w finale"}],
                [{"name": "2", "score": "miejsce w finale"}],
            ],
        )

        self.assertEqual(len(logical), 1)
        self.assertEqual(logical[0].teams, [])

    def test_classic_team_match_is_never_absorbed_by_adjacent_identity(self):
        base = {
            "league": "Test league",
            "track": "Test track",
            "competition": "Team competition",
            "round": "1 runda",
            "capacity": "",
        }
        logical = self.decode_fragments(
            [
                {**base, "home": "Team A", "away": "Team B", "score": "49-41"},
                {**base, "home": "", "away": "", "score": ""},
            ],
            counts=[10, 2],
        )

        self.assertEqual(len(logical), 2)

    def test_canonical_track_punctuation_matches_pwa_signature(self):
        base = {
            "league": "International",
            "competition": "Test Cup",
            "round": "1 runda",
            "capacity": "",
        }
        logical = self.decode_fragments(
            [
                {**base, "track": "King’s-Lynn"},
                {**base, "track": "Kings Lynn"},
            ]
        )

        self.assertEqual(len(logical), 1)

    def test_same_identity_with_different_dates_never_merges(self):
        base = {
            "league": "International",
            "track": "Test track",
            "competition": "Test Cup",
            "round": "1 runda",
            "capacity": "",
        }
        logical = self.decode_fragments(
            [base, base], dates=["2026-08-30", "2026-08-31"]
        )

        self.assertEqual(len(logical), 2)


class ConservativeMatchingTests(unittest.TestCase):
    def assert_high(self, event, calendar, expected_date):
        matched, ambiguous, unmatched = match_events([event], calendar)
        self.assertEqual((len(matched), len(ambiguous), len(unmatched)), (1, 0, 0))
        self.assertEqual(matched[0]["event_date"], expected_date)
        return matched[0]

    def test_polish_league_match(self):
        event = make_event(
            home="Wilki Krosno",
            away="Polonia Bydgoszcz",
            score="49-41",
            league="I Liga",
            track="Krosno",
            competition="",
            round="",
        )
        self.assert_high(
            event,
            [candidate("2026-04-05", "Krosno", "KRO - BYD - Runda 1", "Metalkas 2. Ekstraliga, Część zasadnicza")],
            "2026-04-05",
        )

    def test_sgp_and_sec_rounds(self):
        self.assert_high(make_event(), [candidate()], "2026-08-01")
        sec = make_event(track="Stralsund", competition="SEC", round="Eliminacje")
        self.assert_high(
            sec,
            [candidate("2026-04-25", "Stralsund", "Speedway Euro Championship - eliminacje 2", "SEC")],
            "2026-04-25",
        )

    def test_multi_team_event_assigns_one_date_to_every_fragment(self):
        base = {
            "home": "BYD",
            "away": "",
            "score": "10",
            "league": "Polska",
            "track": "Bydgoszcz",
            "competition": "DMPJ",
            "round": "1 runda",
            "capacity": "",
        }
        second = {**base, "home": "KRO", "score": "8"}
        event = make_event(
            fragments=[base, second],
            teams=[{"name": "BYD", "score": "10"}, {"name": "KRO", "score": "8"}],
            **base,
        )
        matched = self.assert_high(
            event,
            [candidate("2026-05-01", "Bydgoszcz", "BYD - KRO - Runda 1", "DMPJ")],
            "2026-05-01",
        )
        approved = {key: matched["event_date"] for key in matched["mapping_keys"]}
        self.assertEqual(len(approved), 2)
        self.assertEqual(set(approved.values()), {"2026-05-01"})

    def test_same_venue_different_dates_uses_round(self):
        event = make_event(round="7 runda")
        matched = self.assert_high(
            event,
            [
                candidate("2026-08-01", event="FIM SGP of Poland - Lodz - Runda 7"),
                candidate("2026-08-02", event="FIM SGP2 of Poland - Lodz - Runda 8", competition="SGP2"),
            ],
            "2026-08-01",
        )
        self.assertEqual(matched["calendar_candidate"]["candidate_index"], 0)

    def test_explicit_round_conflict_is_never_high(self):
        event = make_event(round="7 runda")
        conflicting = candidate(event="FIM SGP of Poland - Lodz - Runda 1")
        matched, ambiguous, unmatched = match_events([event], [conflicting])
        self.assertEqual(len(matched), 0)
        self.assertEqual(len(ambiguous) + len(unmatched), 1)

    def test_one_calendar_candidate_cannot_update_two_events(self):
        events = [make_event(index=0), make_event(index=1)]
        matched, ambiguous, unmatched = match_events(events, [candidate()])
        self.assertEqual(len(matched), 0)
        self.assertEqual(len(ambiguous), 2)
        self.assertEqual(len(unmatched), 0)
        self.assertTrue(all("multiple WZDB events" in item["reason"] for item in ambiguous))

    def test_unmatched_and_future_candidate_are_not_approved(self):
        unmatched_event = make_event(track="Vojens")
        matched, ambiguous, unmatched = match_events(unmatched_event and [unmatched_event], [candidate()])
        self.assertEqual((len(matched), len(ambiguous), len(unmatched)), (0, 0, 1))

        future = candidate("2026-09-01")
        matched, ambiguous, unmatched = match_events([make_event()], [future])
        self.assertEqual((len(matched), len(ambiguous), len(unmatched)), (0, 1, 0))
        self.assertIn("after the analysis cutoff", ambiguous[0]["reason"])

    def test_diacritics_and_foreign_aliases(self):
        event = make_event(track="Praga", round="2 runda")
        self.assert_high(
            event,
            [candidate("2026-05-23", "Prague", "FIM SGP of Czech Republic - Prague - Runda 2")],
            "2026-05-23",
        )


class EventDateFormatTests(unittest.TestCase):
    def test_mapping_loader_requires_canonical_iso_dates(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "dates.json"
            key = "a" * 64
            path.write_text(f'{{"events":{{"{key}":"2026-08-01"}}}}', encoding="utf-8")
            self.assertEqual(load_event_date_mapping(path), {key: "2026-08-01"})
            path.write_text(f'{{"events":{{"{key}":"01.08.2026"}}}}', encoding="utf-8")
            with self.assertRaises(ValueError):
                load_event_date_mapping(path)

    def test_builder_stores_known_date_once_in_event_reference(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            workbook = Workbook()
            players = workbook.active
            players.title = "Zawodnicy"
            players.append(["Zawodnik", "Narodowość", "Data urodzenia"])
            players.append(["Test Rider", "Polska", None])
            for year in range(2010, 2027):
                sheet = workbook.create_sheet(str(year))
                if year == 2026:
                    sheet.cell(4, 2, "Test Rider")
                    sheet.cell(3, 17, "Data")
                    sheet.cell(4, 17, "05.04.2026")
                    values = ["Wilki Krosno", "Polonia Bydgoszcz", "49-41", "I Liga", "Krosno", "", "", ""]
                    for column, value in zip((7, 8, 9, 10, 11, 12, 13, 15), values):
                        sheet.cell(4, column, value)
            workbook.save(source)

            signature = tuple(event_component(value) for value in values)
            key = event_mapping_key("2026", signature, 0)
            database = build_database(source, "test", "2026-08-17T00:00:00Z", {key: "2026-04-05"})
            repeated = build_database(source, "test", "2026-08-17T00:00:00Z", {key: "2026-04-05"})

        event_ref = database["events"]["2026"][0]
        self.assertEqual(database, repeated)
        self.assertEqual(event_ref[:2], [0, 1])
        self.assertEqual(event_ref[2:4], [1, []])
        self.assertEqual(database["strings"][event_ref[4]], "2026-04-05")
        self.assertEqual(database["strings"].count("2026-04-05"), 1)
        self.assertTrue(all(len(ref) == 2 for ref in database["events"]["2025"]))

        stats = event_date_assignment_stats(database, {key: "2026-04-05"})
        self.assertEqual(stats["logical_events"], 1)
        self.assertEqual(stats["dated_events"], 1)
        self.assertEqual(stats["dated_records"], 1)
        self.assertEqual(stats["ambiguous_events"], 0)

    def test_unknown_mapping_key_does_not_receive_a_date(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            workbook = Workbook()
            players = workbook.active
            players.title = "Zawodnicy"
            players.append(["Zawodnik", "Narodowość", "Data urodzenia"])
            players.append(["Test Rider", "Polska", None])
            for year in range(2010, 2027):
                sheet = workbook.create_sheet(str(year))
                if year == 2026:
                    sheet.cell(4, 2, "Test Rider")
                    sheet.cell(4, 11, "Krosno")
            workbook.save(source)
            database = build_database(
                source,
                "test",
                "2026-08-17T00:00:00Z",
                {"f" * 64: "2026-04-05"},
            )

        self.assertEqual(database["events"]["2026"][0], [0, 1])
        stats = event_date_assignment_stats(database, {"f" * 64: "2026-04-05"})
        self.assertEqual(stats["dated_events"], 0)
        self.assertEqual(stats["unmatched_events"], 1)
        self.assertEqual(stats["stale_mapping_keys"], 1)


class ReviewedSupplementTests(unittest.TestCase):
    def test_aliases_cover_reviewed_location_names(self):
        self.assertEqual(venue_key("Gorican"), venue_key("Donji Kraljevec"))
        self.assertEqual(venue_key("Middlesbrough"), venue_key("Redcar"))
        self.assertEqual(venue_key("Middlesbrough"), venue_key("ECCO Arena"))

    def test_reviewed_high_replaces_unmatched_only_after_identity_check(self):
        event = make_event(index=7, track="Praga", competition="SoN 2", round="Finał")
        matched, ambiguous, unmatched = match_events([event], [])
        decision = {
            "event_id": event.logical_event_index,
            "status": "HIGH",
            "date": "2026-05-22",
            "identity": {field: event.values[field] for field in EVENT_FIELDS},
            "mapping_keys": [fragment.mapping_key for fragment in event.physical],
            "basis": "review fixture",
            "review": "exact identity",
        }
        matched, ambiguous, unmatched = apply_reviewed_supplement(
            [event], matched, ambiguous, unmatched, {"events": [decision]}
        )
        self.assertEqual((len(matched), len(ambiguous), len(unmatched)), (1, 0, 0))
        self.assertEqual(matched[0]["event_date"], "2026-05-22")
        self.assertEqual(matched[0]["calendar_candidate"]["match_kind"], "reviewed_supplement")

        decision["identity"] = {**decision["identity"], "track": "Inny tor"}
        with self.assertRaises(ValueError):
            apply_reviewed_supplement([event], [], [], [event_report_fixture(event)], {"events": [decision]})

    def test_checked_top30_contains_only_high_and_resolves_event_598(self):
        supplement = json.loads(Path("data/event_dates_2026_supplement.json").read_text(encoding="utf-8"))
        self.assertEqual(len(supplement["events"]), 30)
        self.assertTrue(all(item["status"] == "HIGH" for item in supplement["events"]))
        event_598 = next(item for item in supplement["events"] if item["event_id"] == 598)
        self.assertEqual(event_598["date"], "2026-06-16")
        self.assertIn("pełna obsada", event_598["review"])
        verified = [item for item in supplement["events"] if item["original_status"] == "VERIFY_CATEGORY"]
        self.assertEqual({item["event_id"] for item in verified}, {685, 919, 929})


def event_report_fixture(event):
    return {
        "event_index": event.logical_event_index,
        "event_date": "",
    }


if __name__ == "__main__":
    unittest.main()
