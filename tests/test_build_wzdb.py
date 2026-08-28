import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils.datetime import to_excel

from scripts.build_wzdb import (
    StringTable,
    build_database,
    event_component,
    event_date_assignment_stats,
    record_value,
)
from scripts.event_dates import event_mapping_key


class RecordValueTests(unittest.TestCase):
    def test_string_table_reserves_empty_string_at_zero(self) -> None:
        strings = StringTable()

        self.assertEqual(strings.values[0], "")
        self.assertEqual(strings.intern(""), 0)

    def test_numeric_score_decodes_through_string_table(self) -> None:
        strings = StringTable()

        encoded_score = record_value(7, strings)

        self.assertIsInstance(encoded_score, int)
        self.assertEqual(strings.values[encoded_score], "7")

    def test_result_value_variants_are_preserved(self) -> None:
        strings = StringTable()
        values = [0, 1, 10, 14, 7.0, 3.5, "14+2", "3,2,1", "d", "u", "t"]

        decoded = [strings.values[record_value(value, strings)] for value in values]

        self.assertEqual(
            decoded,
            ["0", "1", "10", "14", "7", "3.5", "14+2", "3,2,1", "d", "u", "t"],
        )

    def test_empty_cells_are_null_references(self) -> None:
        strings = StringTable()

        self.assertIsNone(record_value(None, strings))
        self.assertIsNone(record_value("   ", strings))


class UpdateWorkflowTests(unittest.TestCase):
    def test_drive_download_bypasses_cache_and_checks_frequently(self) -> None:
        workflow = Path(".github/workflows/update-db.yml").read_text(encoding="utf-8")

        self.assertIn('cron: "*/15 * * * *"', workflow)
        self.assertIn("Cache-Control: no-cache", workflow)
        self.assertIn("cachebust=$REQUEST_NONCE", workflow)
        self.assertIn("Google Drive source SHA-256", workflow)
        self.assertNotIn("sha256sum data/event_dates_2026.json", workflow)
        self.assertIn("--require-complete-event-dates", workflow)
        self.assertIn("--expect-date-conflicts 0", workflow)


class RecordLayoutTests(unittest.TestCase):
    def test_column_a_is_appended_without_moving_existing_fields_or_event_key(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            workbook = Workbook()
            players = workbook.active
            players.title = "Zawodnicy"
            players.append(["Zawodnik", "Narodowość", "Data urodzenia"])
            players.append(["Test Rider", "Polska", None])
            for year in range(2010, 2027):
                sheet = workbook.create_sheet(str(year))
                if year == 2026:
                    values = {
                        1: 95, 2: "Test Rider", 3: "11+2", 4: "3,3,2*,2,1*",
                        5: "Pole E", 6: "Pole F",
                        7: "Home", 8: "Away", 9: "49-41", 10: "Liga",
                        11: "Krosno", 12: "Rozgrywki", 13: "1 runda",
                        14: 2026, 15: "500cc", 16: "Uwagi", 17: "23.08.2026",
                    }
                    for column, value in values.items():
                        sheet.cell(4, column, value)
                    sheet.cell(5, 2, "Rider Without Number")
                    for column, value in values.items():
                        if column not in (1, 2):
                            sheet.cell(5, column, value)
            workbook.save(source)

            event_values = ["Home", "Away", "49-41", "Liga", "Krosno", "Rozgrywki", "1 runda", "500cc"]
            signature = tuple(event_component(value) for value in event_values)
            mapping_key = event_mapping_key("2026", signature, 0)
            database = build_database(
                source,
                "test",
                "2026-08-20T00:00:00Z",
                {mapping_key: "2026-08-23"},
            )

        record = database["years"]["2026"][0]
        decoded = [None if value is None else database["strings"][value] for value in record[1:]]
        self.assertEqual(len(record), 15)
        self.assertEqual(
            decoded,
            [
                "11+2", "3,3,2*,2,1*", "Pole E", "Pole F", "Home", "Away",
                "49-41", "Liga", "Krosno", "Rozgrywki", "1 runda", "500cc",
                "Uwagi", "95",
            ],
        )
        self.assertNotIn("2026", decoded)
        self.assertIsNone(database["years"]["2026"][1][14])

        event_ref = database["events"]["2026"][0]
        self.assertEqual(event_ref[:2], [0, 2])
        self.assertEqual(database["strings"][event_ref[4]], "2026-08-23")
        self.assertEqual(database["eventDateDiagnostics"], [])

    def test_missing_date_map_keeps_generation_compatible(self) -> None:
        self.assertEqual(record_value(None, StringTable()), None)


class EventDateSourceTests(unittest.TestCase):
    def build_dates(self, values, event_dates=None):
        import tempfile

        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            workbook = Workbook()
            players = workbook.active
            players.title = "Zawodnicy"
            players.append(["Zawodnik", "Narodowość", "Data urodzenia"])
            players.append(["Test Rider", "Polska", None])
            for year in range(2010, 2027):
                sheet = workbook.create_sheet(str(year))
                sheet.cell(3, 17, "Data")
                if year == 2026:
                    for offset, value in enumerate(values):
                        row = 4 + offset
                        sheet.cell(row, 1, offset + 1)
                        sheet.cell(row, 2, "Test Rider")
                        sheet.cell(row, 7, "Home")
                        sheet.cell(row, 8, "Away")
                        sheet.cell(row, 9, "49-41")
                        sheet.cell(row, 10, "Liga")
                        sheet.cell(row, 11, "Krosno")
                        sheet.cell(row, 12, "Rozgrywki")
                        sheet.cell(row, 13, "1 runda")
                        sheet.cell(row, 15, "500cc")
                        sheet.cell(row, 17, value)
            workbook.save(source)
            return build_database(
                source,
                "test",
                "2026-08-23T00:00:00Z",
                event_dates,
            )

    @staticmethod
    def event_date(database):
        event_ref = database["events"]["2026"][0]
        return None if len(event_ref) < 5 else database["strings"][event_ref[4]]

    def test_one_date_is_assigned_to_the_physical_event(self) -> None:
        database = self.build_dates(["23.08.2026"] * 10)

        self.assertEqual(self.event_date(database), "2026-08-23")

    def test_excel_serial_is_normalized(self) -> None:
        database = self.build_dates([to_excel(date(2026, 8, 23))] * 10)

        self.assertEqual(self.event_date(database), "2026-08-23")

    def test_all_empty_dates_leave_event_undated(self) -> None:
        database = self.build_dates([None] * 10)

        self.assertIsNone(self.event_date(database))

    def test_partial_empty_dates_preserve_the_unique_date(self) -> None:
        database = self.build_dates(["23/08/2026"] * 8 + [None, None])

        self.assertEqual(self.event_date(database), "2026-08-23")

    def test_conflicting_dates_are_null_and_diagnostic(self) -> None:
        database = self.build_dates(
            ["23.08.2026"] * 5 + ["24.08.2026"] * 5
        )

        self.assertIsNone(self.event_date(database))
        [diagnostic] = database["eventDateDiagnostics"]
        self.assertEqual(diagnostic["type"], "source_conflict")
        self.assertEqual(diagnostic["source_rows"], [4, 13])
        self.assertEqual(diagnostic["dates"], ["2026-08-23", "2026-08-24"])
        stats = event_date_assignment_stats(database)
        self.assertEqual(stats["ambiguous_events"], 1)
        self.assertEqual(stats["ambiguous_records"], 10)

    def test_pl2_wins_over_the_optional_json_map(self) -> None:
        signature = ("Home", "Away", "49-41", "Liga", "Krosno", "Rozgrywki", "1 runda", "500cc")
        mapping_key = event_mapping_key("2026", signature, 0)
        database = self.build_dates(
            ["23.08.2026"] * 10,
            {mapping_key: "2026-08-24"},
        )

        self.assertEqual(self.event_date(database), "2026-08-23")
        self.assertEqual(database["eventDateDiagnostics"][0]["type"], "json_mismatch")

    def test_build_without_json_uses_column_q(self) -> None:
        database = self.build_dates(["2026-08-23"] * 10, None)
        stats = event_date_assignment_stats(database, {})

        self.assertEqual(self.event_date(database), "2026-08-23")
        self.assertEqual(stats["dated_events"], 1)
        self.assertEqual(stats["dated_records"], 10)
        self.assertEqual(stats["date_map_mapping_keys"], 0)

    def test_logical_merge_preserves_identical_fragment_dates(self) -> None:
        database = self.build_dates(["23.08.2026"] * 10)
        strings = database["strings"]

        def intern(value):
            if value in strings:
                return strings.index(value)
            strings.append(value)
            return len(strings) - 1

        team_a, team_b = intern("Team A"), intern("Team B")
        for row in database["years"]["2026"][:5]:
            row[5], row[6] = team_a, None
        for row in database["years"]["2026"][5:]:
            row[5], row[6] = None, team_b
        date_index = database["events"]["2026"][0][4]
        database["events"]["2026"] = [
            [0, 5, 1, [], date_index],
            [5, 5, 1, [], date_index],
        ]

        stats = event_date_assignment_stats(database)

        self.assertEqual(stats["logical_events"], 1)
        self.assertEqual(stats["dated_events"], 1)
        self.assertEqual(stats["dated_records"], 10)


if __name__ == "__main__":
    unittest.main()
