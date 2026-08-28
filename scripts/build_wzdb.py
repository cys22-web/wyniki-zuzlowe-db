#!/usr/bin/env python3
"""Build a compact Wyniki Zuzlowe v4 database from the source XLSM."""

from __future__ import annotations

import argparse
import gzip
import hashlib
import json
import math
import os
import re
import sys
import tempfile
import unicodedata
from collections import Counter
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils.datetime import WINDOWS_EPOCH, from_excel

try:
    from scripts.event_dates import (
        DEFAULT_EVENT_DATES,
        event_mapping_key,
        load_event_date_mapping,
    )
except ModuleNotFoundError:  # Direct execution: python scripts/build_wzdb.py
    from event_dates import (
        DEFAULT_EVENT_DATES,
        event_mapping_key,
        load_event_date_mapping,
    )


FORMAT_VERSION = 4
FIRST_SEASON = 2010
LAST_SEASON = 2026
PLAYER_SHEET = "Zawodnicy"
DEFAULT_SOURCE_URL = (
    "https://drive.google.com/uc?"
    "id=15p7L2RPKcSMIVzoZXkGDkZZqR5xlBA9K&export=download"
)


class StringTable:
    """Insertion-ordered string table used by records and player metadata."""

    def __init__(self) -> None:
        # Index 0 is the empty string in the format consumed by the app.
        self.values: list[str] = [""]
        self._indices: dict[str, int] = {"": 0}

    def intern(self, value: str) -> int:
        index = self._indices.get(value)
        if index is None:
            index = len(self.values)
            self._indices[value] = index
            self.values.append(value)
        return index


def clean_text(value: Any) -> str:
    """Remove accidental cell padding while retaining the displayed spelling."""
    return " ".join(str(value).split())


def normalize_name(value: Any) -> str:
    """Create the search/matching form used by Wyniki Zuzlowe v4.

    NFD decomposition removes combining accents but intentionally leaves letters
    such as Polish ``ł`` distinct from ``l``.
    """
    folded = clean_text(value).casefold()
    return "".join(
        char
        for char in unicodedata.normalize("NFD", folded)
        if not unicodedata.combining(char)
    )


def birth_date_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return None
    if isinstance(value, str):
        return clean_text(value) or None
    return value


def event_date_value(value: Any, epoch: datetime = WINDOWS_EPOCH) -> str | None:
    """Normalize a source event date to ISO without inventing missing values."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        raise ValueError(f"wartość logiczna nie jest datą: {value!r}")
    if isinstance(value, (int, float)):
        if not math.isfinite(float(value)):
            raise ValueError(f"nieprawidłowy serial Excela: {value!r}")
        parsed = from_excel(value, epoch)
        if isinstance(parsed, datetime):
            return parsed.date().isoformat()
        if isinstance(parsed, date):
            return parsed.isoformat()
        raise ValueError(f"serial Excela nie wskazuje daty: {value!r}")

    text = clean_text(value)
    if not text:
        return None
    if re.fullmatch(r"\d+(?:\.\d+)?", text):
        return event_date_value(float(text), epoch)
    for pattern in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            pass
    raise ValueError(f"nieobsługiwany format daty: {text!r}")


def resolve_event_date(
    values: Iterable[Any], epoch: datetime = WINDOWS_EPOCH
) -> tuple[str | None, list[str], list[str]]:
    """Resolve one physical event date from all its source-row values."""
    dates: set[str] = set()
    invalid: list[str] = []
    for value in values:
        try:
            parsed = event_date_value(value, epoch)
        except (TypeError, ValueError, OverflowError) as error:
            invalid.append(f"{value!r}: {error}")
            continue
        if parsed:
            dates.add(parsed)
    candidates = sorted(dates)
    if invalid or len(candidates) != 1:
        return None, candidates, invalid
    return candidates[0], candidates, invalid


def record_value(value: Any, strings: StringTable) -> int | None:
    """Encode a result cell as a string-table reference used by the app."""
    if value is None:
        return None
    if isinstance(value, str):
        text = clean_text(value)
        return None if not text else strings.intern(text)
    if isinstance(value, datetime):
        return strings.intern(value.isoformat())
    if isinstance(value, (date, time)):
        return strings.intern(value.isoformat())
    if isinstance(value, timedelta):
        value = value.total_seconds()
    if isinstance(value, float) and not math.isfinite(value):
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return strings.intern(str(value))


def event_component(value: Any) -> str:
    """Canonical event-key component resilient to Excel text/number drift."""
    if value is None:
        return ""
    if isinstance(value, str):
        return clean_text(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, timedelta):
        return str(value.total_seconds())
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def generation_input_sha256(paths: Iterable[Path]) -> str:
    """Hash every file that can change the generated WZDB representation."""
    digest = hashlib.sha256()
    for path in paths:
        resolved = path.resolve()
        data = resolved.read_bytes()
        digest.update(resolved.name.encode("utf-8"))
        digest.update(b"\0")
        digest.update(len(data).to_bytes(8, "big"))
        digest.update(data)
    return digest.hexdigest()


def atomic_write(path: Path, data: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".tmp", dir=path.parent
    )
    try:
        with os.fdopen(descriptor, "wb") as output:
            output.write(data)
            output.flush()
            os.fsync(output.fileno())
        os.replace(temporary_name, path)
    except BaseException:
        try:
            os.unlink(temporary_name)
        except FileNotFoundError:
            pass
        raise


def iter_player_rows(sheet: Any) -> Iterable[tuple[Any, Any, Any]]:
    for name, nationality, born in sheet.iter_rows(
        min_row=2, min_col=1, max_col=3, values_only=True
    ):
        if name is not None and clean_text(name):
            yield name, nationality, born


def build_database(
    source_path: Path,
    source_url: str,
    built: str,
    event_dates: dict[str, str] | None = None,
) -> dict[str, Any]:
    workbook = load_workbook(
        source_path,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        expected_sheets = [str(year) for year in range(FIRST_SEASON, LAST_SEASON + 1)]
        missing = [
            name for name in [PLAYER_SHEET, *expected_sheets] if name not in workbook.sheetnames
        ]
        if missing:
            raise ValueError(f"Brak wymaganych arkuszy: {', '.join(missing)}")

        strings = StringTable()
        players: list[list[Any]] = []
        player_by_normalized_name: dict[str, int] = {}

        for raw_name, raw_nationality, raw_born in iter_player_rows(
            workbook[PLAYER_SHEET]
        ):
            name = clean_text(raw_name)
            normalized = normalize_name(name)
            nationality = clean_text(raw_nationality) if raw_nationality is not None else ""
            player_id = len(players)
            players.append(
                [name, strings.intern(nationality), birth_date_value(raw_born), normalized]
            )
            player_by_normalized_name.setdefault(normalized, player_id)

        years: dict[str, list[list[Any]]] = {}
        events: dict[str, list[list[Any]]] = {}
        source_event_rows: dict[str, list[tuple[int, int]]] = {}
        event_date_diagnostics: list[dict[str, Any]] = []
        total_rows = 0
        total_events = 0
        approved_event_dates = event_dates or {}

        for year in range(FIRST_SEASON, LAST_SEASON + 1):
            year_key = str(year)
            records: list[list[Any]] = []
            year_events: list[list[Any]] = []
            year_source_rows: list[tuple[int, int]] = []
            current_event_key: tuple[str, ...] | None = None
            current_mapping_key: str | None = None
            current_source_start: int | None = None
            current_source_end: int | None = None
            current_date_values: list[Any] = []
            event_occurrences: Counter[tuple[str, ...]] = Counter()

            def finalize_event() -> None:
                nonlocal current_mapping_key
                if not year_events or current_mapping_key is None:
                    return
                event_index = len(year_events) - 1
                event_ref = year_events[event_index]
                source_start = current_source_start or 0
                source_end = current_source_end or source_start
                event_date, candidates, invalid = resolve_event_date(
                    current_date_values, workbook.epoch
                )
                if event_date:
                    # WZDB v4 backward-compatible event extension:
                    # [start, count, fragmentCount, teams, eventDateStringIndex]
                    event_ref.extend([1, [], strings.intern(event_date)])
                if len(candidates) > 1:
                    event_date_diagnostics.append(
                        {
                            "type": "source_conflict",
                            "season": int(year_key),
                            "event_index": event_index,
                            "mapping_key": current_mapping_key,
                            "start": event_ref[0],
                            "count": event_ref[1],
                            "source_rows": [source_start, source_end],
                            "dates": candidates,
                        }
                    )
                if invalid:
                    event_date_diagnostics.append(
                        {
                            "type": "invalid_source_date",
                            "season": int(year_key),
                            "event_index": event_index,
                            "mapping_key": current_mapping_key,
                            "start": event_ref[0],
                            "count": event_ref[1],
                            "source_rows": [source_start, source_end],
                            "dates": candidates,
                            "invalid_values": invalid,
                        }
                    )
                mapped_date = approved_event_dates.get(current_mapping_key)
                if mapped_date and mapped_date != event_date:
                    event_date_diagnostics.append(
                        {
                            "type": "json_mismatch",
                            "season": int(year_key),
                            "event_index": event_index,
                            "mapping_key": current_mapping_key,
                            "start": event_ref[0],
                            "count": event_ref[1],
                            "source_rows": [source_start, source_end],
                            "source_date": event_date,
                            "json_date": mapped_date,
                        }
                    )
                year_source_rows.append((source_start, source_end))

            # A:Q is 17 cells. Existing record indices remain unchanged:
            # B is the player lookup, C:M and O:P become row[1]..row[13].
            # Source column A is appended as the optional start number at row[14];
            # N is deliberately omitted, while Q supplies only the event date.
            for source_row, cells in enumerate(
                workbook[year_key].iter_rows(
                    min_row=4, min_col=1, max_col=17, values_only=True
                ),
                start=4,
            ):
                raw_player_name = cells[1]
                if raw_player_name is None or not clean_text(raw_player_name):
                    continue

                player_name = clean_text(raw_player_name)
                normalized = normalize_name(player_name)
                player_id = player_by_normalized_name.get(normalized)
                if player_id is None:
                    player_id = len(players)
                    player_by_normalized_name[normalized] = player_id
                    players.append(
                        [player_name, strings.intern(""), None, normalized]
                    )

                raw_record_values = [*cells[2:13], cells[14], cells[15], cells[0]]
                record = [player_id]
                record.extend(record_value(value, strings) for value in raw_record_values)
                records.append(record)

                # Record indices 5..12 (inclusive) still correspond to source G:M,O.
                raw_event_values = (*cells[6:13], cells[14])
                event_key = tuple(event_component(value) for value in raw_event_values)
                record_index = len(records) - 1
                if event_key == current_event_key:
                    year_events[-1][1] += 1
                    current_source_end = source_row
                    current_date_values.append(cells[16])
                else:
                    finalize_event()
                    occurrence = event_occurrences[event_key]
                    event_occurrences[event_key] += 1
                    current_mapping_key = event_mapping_key(
                        year_key, event_key, occurrence
                    )
                    year_events.append([record_index, 1])
                    current_event_key = event_key
                    current_source_start = source_row
                    current_source_end = source_row
                    current_date_values = [cells[16]]

            finalize_event()

            years[year_key] = records
            events[year_key] = year_events
            source_event_rows[year_key] = year_source_rows
            total_rows += len(records)
            total_events += len(year_events)

        stats = {
            "rows": total_rows,
            "players": len(players),
            "seasons": len(years),
            "from": FIRST_SEASON,
            "to": LAST_SEASON,
            "events": total_events,
        }
        database = {
            "version": FORMAT_VERSION,
            "source": source_url,
            "built": built,
            "strings": strings.values,
            "players": players,
            "years": years,
            "stats": stats,
            "events": events,
            "eventDateDiagnostics": event_date_diagnostics,
        }
        append_logical_event_date_conflicts(database, source_event_rows)
        return database
    finally:
        workbook.close()


def decode_event_ref_date(database: dict[str, Any], season: str, index: int) -> str | None:
    event_ref = database["events"][season][index]
    if len(event_ref) < 5 or event_ref[4] is None:
        return None
    value = event_ref[4]
    if isinstance(value, str):
        return value or None
    return str(database["strings"][value] or "") or None


def decode_logical_events(database: dict[str, Any], season: str) -> list[Any]:
    try:
        from scripts.match_event_dates import decode_events
    except ModuleNotFoundError:  # Direct execution: python scripts/build_wzdb.py
        from match_event_dates import decode_events
    return decode_events(database, season)


def append_logical_event_date_conflicts(
    database: dict[str, Any],
    source_event_rows: dict[str, list[tuple[int, int]]],
) -> None:
    """Report date disagreement introduced only when physical fragments merge."""
    diagnostics = database["eventDateDiagnostics"]
    for season in database["years"]:
        for event in decode_logical_events(database, season):
            dates = sorted(
                {
                    value
                    for item in event.physical
                    if (
                        value := decode_event_ref_date(
                            database, season, item.source_event_index
                        )
                    )
                }
            )
            if len(dates) < 2:
                continue
            physical_indexes = [item.source_event_index for item in event.physical]
            ranges = [source_event_rows[season][index] for index in physical_indexes]
            diagnostics.append(
                {
                    "type": "logical_conflict",
                    "season": int(season),
                    "event_index": event.logical_event_index,
                    "physical_event_indexes": physical_indexes,
                    "mapping_keys": [item.mapping_key for item in event.physical],
                    "start": event.start,
                    "count": event.count,
                    "source_rows": [
                        min(item[0] for item in ranges),
                        max(item[1] for item in ranges),
                    ],
                    "dates": dates,
                }
            )


def event_date_assignment_stats(
    database: dict[str, Any],
    event_dates: dict[str, str] | None = None,
    season: str = "2026",
) -> dict[str, Any]:
    """Audit PL2 date coverage; the optional JSON mapping is comparison-only."""
    event_dates = event_dates or {}

    logical_events = decode_logical_events(database, season)
    source_keys = {
        physical.mapping_key
        for event in logical_events
        for physical in event.physical
    }
    diagnostics = [
        item
        for item in database.get("eventDateDiagnostics", [])
        if str(item.get("season")) == season
    ]
    conflicted_physical = {
        int(item["event_index"])
        for item in diagnostics
        if item.get("type") in {"source_conflict", "invalid_source_date"}
        and "event_index" in item
    }
    dated_events = 0
    ambiguous_events = 0
    unmatched_events = 0
    dated_records = 0
    ambiguous_records = 0
    unmatched_records = 0
    for event in logical_events:
        fragment_dates = [
            decode_event_ref_date(database, season, item.source_event_index)
            for item in event.physical
        ]
        known_dates = {value for value in fragment_dates if value}
        has_source_conflict = any(
            item.source_event_index in conflicted_physical for item in event.physical
        )
        if len(known_dates) > 1 or has_source_conflict:
            ambiguous_events += 1
            ambiguous_records += event.count
        elif len(known_dates) == 1:
            dated_events += 1
            dated_records += event.count
        else:
            unmatched_events += 1
            unmatched_records += event.count
    physical_dates = [
        decode_event_ref_date(database, season, index)
        for index in range(len(database["events"][season]))
    ]
    json_matches = 0
    for event in logical_events:
        for item in event.physical:
            source_date = decode_event_ref_date(
                database, season, item.source_event_index
            )
            if source_date and event_dates.get(item.mapping_key) == source_date:
                json_matches += 1
    source_conflicts = sum(
        item.get("type") in {"source_conflict", "invalid_source_date", "logical_conflict"}
        for item in diagnostics
    )
    json_conflicts = sum(item.get("type") == "json_mismatch" for item in diagnostics)
    return {
        "season": int(season),
        "logical_events": len(logical_events),
        "records": len(database["years"][season]),
        "source_mapping_keys": len(source_keys),
        "date_map_mapping_keys": len(event_dates),
        "matching_mapping_keys": len(source_keys & set(event_dates)),
        "stale_mapping_keys": len(set(event_dates) - source_keys),
        "matching_source_and_json_dates": json_matches,
        "physical_events": len(physical_dates),
        "dated_physical_events": sum(bool(value) for value in physical_dates),
        "undated_physical_events": sum(not value for value in physical_dates),
        "dated_events": dated_events,
        "ambiguous_events": ambiguous_events,
        "ambiguous": ambiguous_events,
        "unmatched_events": unmatched_events,
        "events_with_date": dated_events,
        "events_without_date": ambiguous_events + unmatched_events,
        "dated_records": dated_records,
        "ambiguous_records": ambiguous_records,
        "unmatched_records": unmatched_records,
        "records_with_date": dated_records,
        "records_without_date": ambiguous_records + unmatched_records,
        "source_conflicts": source_conflicts,
        "json_conflicts": json_conflicts,
        "conflicts": source_conflicts + json_conflicts,
    }


def validate_expectations(stats: dict[str, int], args: argparse.Namespace) -> None:
    expectations = {
        "rows": args.expect_rows,
        "players": args.expect_players,
        "seasons": args.expect_seasons,
        "from": args.expect_from,
        "to": args.expect_to,
        "events": args.expect_events,
    }
    mismatches = [
        f"{key}: otrzymano {stats[key]}, oczekiwano {expected}"
        for key, expected in expectations.items()
        if expected is not None and stats[key] != expected
    ]
    if mismatches:
        raise ValueError("Niezgodne statystyki: " + "; ".join(mismatches))


def validate_date_expectations(stats: dict[str, Any], args: argparse.Namespace) -> None:
    mismatches: list[str] = []
    if args.expect_date_conflicts is not None and stats["conflicts"] != args.expect_date_conflicts:
        mismatches.append(
            f"conflicts: otrzymano {stats['conflicts']}, "
            f"oczekiwano {args.expect_date_conflicts}"
        )
    if args.require_complete_event_dates:
        if stats["events_without_date"]:
            mismatches.append(
                f"wydarzenia bez jednoznacznej daty: {stats['events_without_date']}"
            )
        if stats["records_without_date"]:
            mismatches.append(
                f"rekordy bez jednoznacznej daty: {stats['records_without_date']}"
            )
    if mismatches:
        raise ValueError("Niezgodna walidacja dat: " + "; ".join(mismatches))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source_xlsm", type=Path, help="Ścieżka do źródłowego PL2.xlsm")
    parser.add_argument("--output", type=Path, default=Path("db/latest.wzdb"))
    parser.add_argument("--version-file", type=Path, default=Path("db/version.json"))
    parser.add_argument("--source-url", default=DEFAULT_SOURCE_URL)
    parser.add_argument(
        "--event-dates",
        type=Path,
        default=DEFAULT_EVENT_DATES,
        help=(
            "Opcjonalna historyczna mapa dat używana wyłącznie do porównania z Q/Data; "
            "jej brak nie blokuje budowy"
        ),
    )
    parser.add_argument(
        "--source-modified",
        help="Wartość Last-Modified zwrócona przy pobieraniu źródła z Google Drive",
    )
    parser.add_argument("--expect-rows", type=int)
    parser.add_argument("--expect-players", type=int)
    parser.add_argument("--expect-seasons", type=int)
    parser.add_argument("--expect-from", type=int)
    parser.add_argument("--expect-to", type=int)
    parser.add_argument("--expect-events", type=int)
    parser.add_argument("--expect-date-conflicts", type=int)
    parser.add_argument(
        "--require-complete-event-dates",
        action="store_true",
        help="Przerwij build, jeśli sezon datowania zawiera wydarzenia bez daty",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    source_path = args.source_xlsm.resolve()
    if not source_path.is_file():
        raise FileNotFoundError(f"Nie znaleziono pliku źródłowego: {source_path}")

    built = datetime.now(timezone.utc).isoformat(timespec="seconds").replace(
        "+00:00", "Z"
    )
    source_sha256 = sha256_file(source_path)
    event_dates_path = args.event_dates.resolve()
    event_dates = load_event_date_mapping(event_dates_path)
    generator_inputs = [
        Path(__file__).resolve(),
        Path(__file__).with_name("event_dates.py"),
        Path(__file__).with_name("match_event_dates.py"),
    ]
    generator_sha256 = generation_input_sha256(generator_inputs)
    database = build_database(source_path, args.source_url, built, event_dates)
    date_stats = event_date_assignment_stats(database, event_dates)
    database["dateStats"] = date_stats
    validate_expectations(database["stats"], args)

    for diagnostic in database.get("eventDateDiagnostics", []):
        rows = diagnostic.get("source_rows", ["?", "?"])
        print(
            "KONFLIKT DATY "
            f"[{diagnostic.get('type')}] sezon={diagnostic.get('season')} "
            f"event={diagnostic.get('event_index')} "
            f"wiersze={rows[0]}-{rows[1]} "
            f"daty={diagnostic.get('dates', [])} "
            f"PL2={diagnostic.get('source_date')} "
            f"JSON={diagnostic.get('json_date')}",
            file=sys.stderr,
        )
    validate_date_expectations(date_stats, args)

    json_bytes = json.dumps(
        database, ensure_ascii=False, separators=(",", ":")
    ).encode("utf-8")
    wzdb_bytes = gzip.compress(json_bytes, compresslevel=9, mtime=0)
    atomic_write(args.output, wzdb_bytes)
    wzdb_sha256 = hashlib.sha256(wzdb_bytes).hexdigest()

    version = {
        "version": FORMAT_VERSION,
        "source": args.source_url,
        "source_sha256": source_sha256,
        "source_hash": source_sha256[:12],
        "generator_sha256": generator_sha256,
        "event_date_source": "PL2.xlsm:Q/Data",
        "date_map_sha256": (
            sha256_file(event_dates_path) if event_dates_path.is_file() else None
        ),
        # Retain the earlier field for consumers that already inspected it.
        "event_dates_sha256": (
            sha256_file(event_dates_path) if event_dates_path.is_file() else None
        ),
        "dated_event_fragments": date_stats["dated_physical_events"],
        "date_stats": date_stats,
        "wzdb_sha256": wzdb_sha256,
        "built": built,
        "stats": database["stats"],
    }
    if args.source_modified:
        version["source_modified"] = args.source_modified
    atomic_write(
        args.version_file,
        (json.dumps(version, ensure_ascii=False, indent=2) + "\n").encode("utf-8"),
    )

    print(json.dumps(version, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
